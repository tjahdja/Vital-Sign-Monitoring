import os
from openpyxl import Workbook, load_workbook

class Dbase:
    def __init__(self):
        super().__init__()
        self.MainDir = os.getcwd()+'\Database'
        print(self.MainDir)

    def create(self, folder):
        self.folder = folder
        if not os.path.exists(f'{self.MainDir}/{self.folder}'):
            os.mkdir(f'{self.MainDir}/{self.folder}')
            print('folder created')
        else:
            print('folder is already exist')

    def WFile(self,fname):
        if os.path.isfile(f'{self.MainDir}/{self.folder}/{fname}.xlsx'):
            self.wb = load_workbook(f'{self.MainDir}/{self.folder}/{fname}.xlsx', read_only=False)

        else:
            self.wb = Workbook()

        self.sheet = self.wb.active
        self.i = 2
        c1 = self.sheet.cell(row=1, column=1)
        c1.value ='Tanggal'

        c2 = self.sheet.cell(row=1, column=2)
        c2.value ='Detak Jantung'

        c3 = self.sheet.cell(row=1, column=3)
        c3.value ='Kadara Oksigen'

        c4 = self.sheet.cell(row=1, column=4)
        c4.value = 'Status'

    def Fillinit(self, paramA, paramB, stat, tgl):
        c1 = self.sheet.cell(row=self.i, column=1)
        c1.value = f'{tgl}'

        c2 = self.sheet.cell(row=self.i, column=2)
        c2.value = f'{paramA}'

        c3 = self.sheet.cell(row=self.i, column=3)
        c3.value = f'{paramB}'

        c4 = self.sheet.cell(row=self.i, column=4)
        c4.value = f'{stat}'

        self.i = self.i+1

    def saveit(self,fname):
        self.wb.save(f'{self.MainDir}/{self.folder}/{fname}.xlsx')
